from __future__ import annotations

import argparse
import html as html_lib
import json
import math
import re
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path.cwd()
RUN_DATE = datetime.now(ZoneInfo("Asia/Shanghai")).date().isoformat()
RUN_TS = datetime.now(ZoneInfo("Asia/Shanghai")).isoformat(timespec="seconds")
UA = "investment-sourcing-agent/0.1 (+Codex daily_run)"
SKILL_PATH = ROOT / "skills" / "investment-sourcing-agent" / "investment-sourcing-agent.skill.md"


def load_runner_policy() -> dict:
    text = SKILL_PATH.read_text(encoding="utf-8")
    match = re.search(r"```json runner_policy\s*(\{.*?\})\s*```", text, re.S)
    if not match:
        raise RuntimeError(f"runner_policy JSON block not found in {SKILL_PATH}")
    return json.loads(match.group(1))


POLICY = load_runner_policy()
CAP = int(POLICY["max_qualified_items_per_source"])
TOP_LIMIT = int(POLICY["top_limit"])
SOURCE_QUERIES = POLICY["source_queries"]
SOURCE_THRESHOLDS = POLICY["source_thresholds"]
OUTREACH_POLICY = POLICY["outreach"]
GITHUB_POLICY = SOURCE_THRESHOLDS["github"]
HN_POLICY = SOURCE_THRESHOLDS["hacker_news"]
ARXIV_POLICY = SOURCE_THRESHOLDS["arxiv"]
YC_POLICY = SOURCE_THRESHOLDS["yc"]


def build_filter_descriptions() -> dict:
    allowed_batches = ", ".join(YC_POLICY["allowed_batches"])
    return {
        "YC": f"Active/Early, batch in {allowed_batches}, team <= {YC_POLICY['max_team_size']}, theme tags >= {YC_POLICY['min_theme_tags']}",
        "GitHub": (
            f"stars >= {GITHUB_POLICY['min_stars']}, or stars/day >= {GITHUB_POLICY['min_stars_per_day']} "
            f"with stars >= {GITHUB_POLICY['velocity_min_stars']}, or repo <= {GITHUB_POLICY['new_repo_days']}d "
            f"with stars >= {GITHUB_POLICY['new_repo_min_stars']}"
        ),
        "arXiv": f"submitted <= {ARXIV_POLICY['max_age_days']}d and theme tags >= {ARXIV_POLICY['min_theme_tags']}",
        "Hacker News": (
            f"external URL and (points >= {HN_POLICY['min_points']} or comments >= {HN_POLICY['min_comments']} "
            f"or Launch/Show HN points >= {HN_POLICY['launch_min_points']})"
        ),
    }


FILTER_DESCRIPTIONS = build_filter_descriptions()

DATA_DIR = ROOT / "data"
RAW_DIR = DATA_DIR / "raw"
REPORT_DIR = ROOT / "reports" / "daily"
MASTER_PATH = DATA_DIR / "investment_sourcing_master.xlsx"
RAW_PATH = RAW_DIR / f"{RUN_DATE}-sourcing-run.json"
REPORT_PATH = REPORT_DIR / f"{RUN_DATE}-sourcing-digest.md"

COLUMNS = [
    "first_seen_date",
    "last_seen_date",
    "source",
    "type",
    "name_or_title",
    "url",
    "score",
    "tier",
    "outreach_status",
    "founder_or_contact",
    "china_signal",
    "founder_background_signal",
    "theme_tags",
    "one_line_summary",
    "evidence",
    "markdown_anchor",
    "notes",
]

THEMES = list(POLICY["theme_terms"].items())
ELITE_TERMS = POLICY["background_terms"]
CHINA_TERMS = POLICY["china_terms"]
CHINESE_NAME_HINTS = set(POLICY["romanized_chinese_name_hints"])
EMAIL_RE = re.compile(r"(?<![\w.])([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})(?!\w)", re.I)
ATOM = "{http://www.w3.org/2005/Atom}"


@dataclass
class Item:
    source: str
    type: str
    name: str
    url: str
    raw_text: str
    published_at: str | None
    metadata: dict
    fetched_at: str


@dataclass
class Scored:
    item: Item
    score: int
    tier: str
    outreach_status: str
    founder_or_contact: str
    china_signal: str
    founder_background_signal: str
    theme_tags: list[str]
    one_line_summary: str
    evidence: list[str]
    markdown_anchor: str
    why: str
    next_action: str
    contact_paths: list[str]
    english_subject: str
    english_body: str
    chinese_subject: str
    chinese_body: str
    breakdown: dict


def clean(value: object, limit: int | None = None) -> str:
    text = html_lib.unescape(str(value or ""))
    text = re.sub(r"</?[A-Za-z][^>]*>", " ", text)
    text = " ".join(text.replace("\r", " ").replace("\n", " ").split())
    if limit and len(text) > limit:
        return text[: limit - 1].rstrip() + "..."
    return text


def http_text(url: str, failures: list[dict], source: str, action: str, *, headers: dict | None = None, data: bytes | None = None) -> str | None:
    req_headers = {"User-Agent": UA}
    if headers:
        req_headers.update(headers)
    try:
        with urllib.request.urlopen(urllib.request.Request(url, data=data, headers=req_headers), timeout=30) as resp:
            return resp.read().decode(resp.headers.get_content_charset() or "utf-8", "replace")
    except Exception as exc:
        failures.append(
            {
                "source": source,
                "action": action,
                "url": url,
                "time": RUN_TS,
                "reason": f"{type(exc).__name__}: {exc}",
                "status": "fetch_failed",
                "retry": "retry source fetch later or reduce query scope",
            }
        )
        return None


def optional_http_text(url: str, *, headers: dict | None = None, data: bytes | None = None) -> str | None:
    req_headers = {"User-Agent": UA}
    if headers:
        req_headers.update(headers)
    try:
        with urllib.request.urlopen(urllib.request.Request(url, data=data, headers=req_headers), timeout=20) as resp:
            return resp.read().decode(resp.headers.get_content_charset() or "utf-8", "replace")
    except Exception:
        return None


def unique(values):
    seen = set()
    out = []
    for value in values:
        if value is None:
            continue
        key = clean(value) if isinstance(value, str) else json.dumps(value, ensure_ascii=False, sort_keys=True)
        if key and key not in seen:
            seen.add(key)
            out.append(value)
    return out


def extract_emails(text: str) -> list[str]:
    return unique([m.group(1).lower() for m in EMAIL_RE.finditer(text or "")])


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")[:60] or "item"


def theme_tags(text: str) -> list[str]:
    low = text.lower()
    return [tag for tag, words in THEMES if any(word in low for word in words)]


def freshness(date_text: str | None, max_points: int) -> int:
    if not date_text:
        return 0
    try:
        if str(date_text).isdigit():
            dt = datetime.fromtimestamp(int(date_text), timezone.utc)
        else:
            dt = datetime.fromisoformat(str(date_text).replace("Z", "+00:00"))
        days = (datetime.now(timezone.utc) - dt.astimezone(timezone.utc)).days
    except Exception:
        return 0
    if days <= 3:
        return max_points
    if days <= 14:
        return int(max_points * 0.8)
    if days <= 45:
        return int(max_points * 0.45)
    return int(max_points * 0.15)


def parse_datetime(date_text: str | None) -> datetime | None:
    if not date_text:
        return None
    try:
        if str(date_text).isdigit():
            return datetime.fromtimestamp(int(date_text), timezone.utc)
        return datetime.fromisoformat(str(date_text).replace("Z", "+00:00")).astimezone(timezone.utc)
    except Exception:
        return None


def age_days(date_text: str | None) -> int | None:
    parsed = parse_datetime(date_text)
    if not parsed:
        return None
    return max(0, (datetime.now(timezone.utc) - parsed).days)


def tier(score: int) -> str:
    if score >= 85:
        return "A"
    if score >= 70:
        return "B"
    if score >= 55:
        return "C"
    return "Watch"


def china_signal(text: str, names: list[object]) -> str:
    low = (text or "").lower()
    if any(term in low for term in CHINA_TERMS):
        return "中等置信度：公开文本出现 China/cross-border 等相关词"
    name_words = set(re.findall(r"[a-z]+", " ".join(clean(name) for name in names if name).lower()))
    hints = sorted(name_words & CHINESE_NAME_HINTS)
    if hints:
        return "低置信度：仅罗马拼音姓名启发式，未验证（" + ", ".join(hints[:3]) + "）"
    if any("\u4e00" <= ch <= "\u9fff" for ch in text or ""):
        return "中等置信度：公开文本出现中文"
    return "未知：未观察到公开中国/华人相关信号"


def background_signal(text: str, founders: list[dict], owner: dict, authors: list[str]) -> str:
    chunks = []
    for founder in founders or []:
        name = clean(founder.get("full_name"))
        bio = clean(founder.get("founder_bio"), 220)
        chunks.append(clean(f"{name}: {bio}" if bio else name))
    owner_bits = [owner.get("name"), owner.get("company"), owner.get("bio"), owner.get("blog"), owner.get("email")]
    owner_text = clean(" | ".join(clean(bit) for bit in owner_bits if bit), 260)
    if owner_text:
        chunks.append(owner_text)
    if authors:
        chunks.append("Authors: " + ", ".join(clean(author) for author in authors[:8] if author))
    if not chunks:
        return "抓取到的公开元数据中未发现"
    signal = clean(" ; ".join(chunks), 420)
    hits = [term for term in ELITE_TERMS if term in signal.lower()]
    return signal + (" | 强信号词：" + ", ".join(hits[:6]) if hits else "")


def passes_yc_filter(company: dict, text: str) -> tuple[bool, str]:
    tags = theme_tags(text)
    batch = company.get("batch_name") or company.get("batch")
    status = clean(company.get("status") or company.get("ycdc_status")).lower()
    stage = clean(company.get("stage")).lower()
    team_size = int(company.get("team_size") or 0)
    active = (not status or status == "active") if YC_POLICY.get("require_active", True) else True
    early = (not stage or stage == "early") if YC_POLICY.get("require_early", True) else True
    passed = (
        active
        and early
        and batch in set(YC_POLICY["allowed_batches"])
        and 0 < team_size <= int(YC_POLICY["max_team_size"])
        and len(tags) >= int(YC_POLICY["min_theme_tags"])
    )
    reason = f"batch={batch}; status={status or 'unknown'}; stage={stage or 'unknown'}; team={team_size or 'unknown'}; themes={','.join(tags) or 'none'}"
    return passed, reason


def passes_github_filter(repo: dict) -> tuple[bool, str]:
    stars = int(repo.get("stargazers_count") or 0)
    days = age_days(repo.get("created_at"))
    stars_per_day = stars / max(days or 1, 1)
    passed = (
        stars >= int(GITHUB_POLICY["min_stars"])
        or (stars >= int(GITHUB_POLICY["velocity_min_stars"]) and stars_per_day >= float(GITHUB_POLICY["min_stars_per_day"]))
        or ((days or 10**9) <= int(GITHUB_POLICY["new_repo_days"]) and stars >= int(GITHUB_POLICY["new_repo_min_stars"]))
    )
    return passed, f"stars={stars}; age_days={days if days is not None else 'unknown'}; stars_per_day={stars_per_day:.2f}"


def passes_arxiv_filter(title: str, summary: str, published: str | None) -> tuple[bool, str]:
    tags = theme_tags(f"{title} {summary}")
    days = age_days(published)
    passed = days is not None and days <= int(ARXIV_POLICY["max_age_days"]) and len(tags) >= int(ARXIV_POLICY["min_theme_tags"])
    return passed, f"age_days={days if days is not None else 'unknown'}; themes={','.join(tags) or 'none'}"


def passes_hn_filter(hit: dict, target_url: str) -> tuple[bool, str]:
    points = int(hit.get("points") or 0)
    comments = int(hit.get("num_comments") or 0)
    title = clean(hit.get("title") or hit.get("story_title"))
    external = bool(target_url and "news.ycombinator.com/item" not in target_url)
    launch = title.lower().startswith(("launch hn", "show hn"))
    external_ok = external if HN_POLICY.get("require_external_url", True) else True
    passed = external_ok and (
        points >= int(HN_POLICY["min_points"])
        or comments >= int(HN_POLICY["min_comments"])
        or (launch and points >= int(HN_POLICY["launch_min_points"]))
    )
    return passed, f"points={points}; comments={comments}; external={external}; launch_or_show={launch}"


def has_china_signal(signal: str) -> bool:
    return not signal.startswith("unknown:") and not signal.startswith("未知")


def passes_outreach_filter(score: int, contacts: list[str], signal: str) -> bool:
    contact_ok = bool(contacts) if OUTREACH_POLICY.get("require_contact_path", True) else True
    china_ok = has_china_signal(signal) if OUTREACH_POLICY.get("require_suspected_china_signal", True) else True
    return score >= int(OUTREACH_POLICY["min_score"]) and contact_ok and china_ok


def fetch_yc(failures: list[dict]) -> list[Item]:
    app = "45BWZJ1SGC"
    key = "NzllNTY5MzJiZGM2OTY2ZTQwMDEzOTNhYWZiZGRjODlhYzVkNjBmOGRjNzJiMWM4ZTU0ZDlhYTZjOTJiMjlhMWFuYWx5dGljc1RhZ3M9eWNkYyZyZXN0cmljdEluZGljZXM9WUNDb21wYW55X3Byb2R1Y3Rpb24lMkNZQ0NvbXBhbnlfQnlfTGF1bmNoX0RhdGVfcHJvZHVjdGlvbiZ0YWdGaWx0ZXJzPSU1QiUyMnljZGNfcHVibGljJTIyJTVE"
    hits = {}
    for query in SOURCE_QUERIES["yc"]:
        params = urllib.parse.urlencode({"query": query, "hitsPerPage": 6, "page": 0})
        url = f"https://{app.lower()}-dsn.algolia.net/1/indexes/YCCompany_By_Launch_Date_production/query"
        text = http_text(
            url,
            failures,
            "YC",
            "algolia_query",
            data=json.dumps({"params": params}).encode(),
            headers={"X-Algolia-Application-Id": app, "X-Algolia-API-Key": key, "Content-Type": "application/json"},
        )
        if text:
            for hit in json.loads(text).get("hits", []):
                hits[hit.get("objectID") or hit.get("slug") or hit.get("name")] = hit
    selected = sorted(hits.values(), key=lambda hit: hit.get("launched_at") or 0, reverse=True)
    items = []
    for hit in selected:
        slug = hit.get("slug")
        detail = {}
        if slug:
            detail_url = f"https://www.ycombinator.com/companies/{slug}"
            detail_html = http_text(detail_url, failures, "YC", "company_detail")
            if detail_html:
                match = re.search(r'data-page="([^"]+)"', detail_html)
                if match:
                    try:
                        detail = json.loads(html_lib.unescape(match.group(1))).get("props", {}).get("company", {}) or {}
                    except Exception as exc:
                        failures.append({"source": "YC", "action": "parse_company_detail", "url": detail_url, "time": RUN_TS, "reason": repr(exc), "status": "fetch_failed", "retry": "retry parsing company page"})
            time.sleep(0.08)
        company = {**hit, **detail}
        founders = company.get("founders") or []
        url = company.get("ycdc_url") or (f"https://www.ycombinator.com/companies/{slug}" if slug else "")
        published = None
        if company.get("launched_at"):
            published = datetime.fromtimestamp(int(company["launched_at"]), timezone.utc).isoformat()
        raw = "\n".join(
            [
                clean(company.get("name")),
                clean(company.get("one_liner")),
                clean(company.get("long_description")),
                "Tags: " + ", ".join(company.get("tags") or []),
                "Batch: " + clean(company.get("batch_name") or company.get("batch")),
                "Founders: " + ", ".join(clean(f.get("full_name")) for f in founders if f.get("full_name")),
                " ".join(clean(f.get("founder_bio")) for f in founders if f.get("founder_bio")),
            ]
        )
        metadata = {
            "one_liner": company.get("one_liner"),
            "description": company.get("long_description"),
            "batch": company.get("batch_name") or company.get("batch"),
            "stage": company.get("stage"),
            "team_size": company.get("team_size"),
            "year_founded": company.get("year_founded"),
            "website": company.get("website"),
            "yc_url": url,
            "twitter_url": company.get("twitter_url"),
            "linkedin_url": company.get("linkedin_url"),
            "github_url": company.get("github_url"),
            "tags": company.get("tags") or [],
            "industries": company.get("industries") or [],
            "founders": founders,
            "source_endpoint": "YC public Algolia index + public company page",
        }
        passed, filter_reason = passes_yc_filter(company, raw)
        metadata["filter_reason"] = filter_reason
        if company.get("name") and url:
            if passed:
                items.append(Item("YC", "company", clean(company.get("name")), url, raw, published, metadata, RUN_TS))
    return items[:CAP]


def fetch_github(failures: list[dict]) -> list[Item]:
    repos = {}
    for query in SOURCE_QUERIES["github"]:
        url = "https://api.github.com/search/repositories?" + urllib.parse.urlencode({"q": query, "sort": "updated", "order": "desc", "per_page": CAP})
        text = http_text(url, failures, "GitHub", "repo_search", headers={"Accept": "application/vnd.github+json"})
        if text:
            for repo in json.loads(text).get("items", []):
                repos[repo.get("full_name") or repo.get("html_url")] = repo
    selected = []
    for repo in repos.values():
        passed, reason = passes_github_filter(repo)
        repo["_filter_reason"] = reason
        if passed:
            selected.append(repo)
    selected = sorted(
        selected,
        key=lambda repo: (
            repo.get("stargazers_count") or 0,
            (repo.get("stargazers_count") or 0) / max(age_days(repo.get("created_at")) or 1, 1),
            repo.get("pushed_at") or "",
        ),
        reverse=True,
    )[:CAP]
    items = []
    for repo in selected:
        owner_profile = {}
        owner_url = repo.get("owner", {}).get("url")
        if owner_url:
            text = optional_http_text(owner_url, headers={"Accept": "application/vnd.github+json"})
            if text:
                owner_profile = json.loads(text)
            time.sleep(0.05)
        full_name = repo.get("full_name") or repo.get("name") or "untitled repository"
        raw = "\n".join(
            [
                full_name,
                clean(repo.get("description")),
                clean(repo.get("homepage")),
                "Topics: " + ", ".join(repo.get("topics") or []),
                "Language: " + clean(repo.get("language")),
                "Owner: " + clean(owner_profile.get("name") or repo.get("owner", {}).get("login")),
                "Owner bio: " + clean(owner_profile.get("bio")),
                "Owner company: " + clean(owner_profile.get("company")),
            ]
        )
        metadata = {
            "description": repo.get("description"),
            "stars": repo.get("stargazers_count"),
            "forks": repo.get("forks_count"),
            "language": repo.get("language"),
            "topics": repo.get("topics") or [],
            "homepage": repo.get("homepage"),
            "html_url": repo.get("html_url"),
            "owner": repo.get("owner", {}).get("login"),
            "owner_html_url": repo.get("owner", {}).get("html_url"),
            "owner_profile": owner_profile,
            "pushed_at": repo.get("pushed_at"),
            "created_at": repo.get("created_at"),
            "filter_reason": repo.get("_filter_reason"),
            "source_endpoint": "GitHub public Search/Users API",
        }
        items.append(Item("GitHub", "repo", full_name, repo.get("html_url") or "", raw, repo.get("pushed_at"), metadata, RUN_TS))
    return items[:CAP]


def fetch_arxiv(failures: list[dict]) -> list[Item]:
    query = SOURCE_QUERIES["arxiv"]
    url = "https://export.arxiv.org/api/query?" + urllib.parse.urlencode({"search_query": query, "sortBy": "submittedDate", "sortOrder": "descending", "max_results": CAP})
    text = http_text(url, failures, "arXiv", "api_query")
    if not text:
        return []
    root = ET.fromstring(text)
    items = []
    for entry in root.findall(ATOM + "entry"):
        title = clean(entry.findtext(ATOM + "title"))
        summary = clean(entry.findtext(ATOM + "summary"))
        paper_url = clean(entry.findtext(ATOM + "id"))
        authors = [clean(a.findtext(ATOM + "name")) for a in entry.findall(ATOM + "author") if clean(a.findtext(ATOM + "name"))]
        cats = [cat.attrib.get("term") for cat in entry.findall(ATOM + "category") if cat.attrib.get("term")]
        raw = "\n".join([title, summary, "Authors: " + ", ".join(authors), "Categories: " + ", ".join(cats)])
        published = clean(entry.findtext(ATOM + "published")) or None
        passed, filter_reason = passes_arxiv_filter(title, summary, published)
        if title and paper_url and passed:
            items.append(
                Item(
                    "arXiv",
                    "paper",
                    title,
                    paper_url,
                    raw,
                    published,
                    {"authors": authors, "categories": cats, "filter_reason": filter_reason, "source_endpoint": "arXiv public API"},
                    RUN_TS,
                )
            )
    return items[:CAP]


def fetch_hn(failures: list[dict]) -> list[Item]:
    since = int((datetime.now(timezone.utc) - timedelta(days=14)).timestamp())
    hits = {}
    for query in SOURCE_QUERIES["hacker_news"]:
        params = {"query": query, "tags": "story", "hitsPerPage": CAP, "numericFilters": "created_at_i>=" + str(since)}
        url = "https://hn.algolia.com/api/v1/search_by_date?" + urllib.parse.urlencode(params)
        text = http_text(url, failures, "Hacker News", "algolia_search")
        if not text:
            continue
        for hit in json.loads(text).get("hits", []):
            title = clean(hit.get("title") or hit.get("story_title"))
            target_url = hit.get("url") or hit.get("story_url") or ("https://news.ycombinator.com/item?id=" + str(hit.get("objectID")))
            combined = f"{title} {hit.get('story_text') or ''} {target_url}".lower()
            passed, filter_reason = passes_hn_filter(hit, target_url)
            if passed and (theme_tags(combined) or any(word in combined for word in ["ai", "llm", "robot", "agent", "inference", "mcp"])):
                hit["_filter_reason"] = filter_reason
                hits[hit.get("objectID") or target_url] = hit
    selected = sorted(hits.values(), key=lambda hit: ((hit.get("points") or 0), hit.get("created_at_i") or 0), reverse=True)[:CAP]
    items = []
    for hit in selected:
        title = clean(hit.get("title") or hit.get("story_title") or "HN story")
        target_url = hit.get("url") or hit.get("story_url") or ("https://news.ycombinator.com/item?id=" + str(hit.get("objectID")))
        hn_url = "https://news.ycombinator.com/item?id=" + str(hit.get("objectID"))
        raw = "\n".join([title, clean(hit.get("story_text") or hit.get("comment_text")), target_url, "HN: " + hn_url, "Author: " + clean(hit.get("author"))])
        metadata = {"points": hit.get("points"), "num_comments": hit.get("num_comments"), "author": hit.get("author"), "hn_url": hn_url, "created_at": hit.get("created_at"), "filter_reason": hit.get("_filter_reason"), "source_endpoint": "HN Algolia Search API"}
        items.append(Item("Hacker News", "post", title, target_url, raw, hit.get("created_at"), metadata, RUN_TS))
    return items[:CAP]


def score_item(item: Item) -> Scored:
    text = f"{item.name} {item.raw_text} {json.dumps(item.metadata, ensure_ascii=False)}"
    low = text.lower()
    themes = theme_tags(text)
    founders = item.metadata.get("founders") or []
    owner = item.metadata.get("owner_profile") or {}
    authors = item.metadata.get("authors") or []
    contacts = extract_emails(text)
    for key in ["website", "homepage", "html_url", "owner_html_url", "twitter_url", "linkedin_url", "github_url", "blog", "yc_url"]:
        value = item.metadata.get(key)
        if isinstance(value, str) and value.startswith("http"):
            contacts.append(value)
    for founder in founders:
        for key in ["twitter_url", "linkedin_url"]:
            value = founder.get(key)
            if isinstance(value, str) and value.startswith("http"):
                contacts.append(value)
    contacts = unique(contacts)
    bg = background_signal(text, founders, owner, authors)
    names = [f.get("full_name") for f in founders] + authors + [owner.get("name"), owner.get("login")]
    china = china_signal(text, names)
    evidence = unique([item.url, item.metadata.get("website"), item.metadata.get("homepage"), item.metadata.get("yc_url"), item.metadata.get("html_url")])
    one_line = clean(item.metadata.get("one_liner") or item.metadata.get("description") or item.raw_text, 180)

    if item.type == "company":
        founder_pts = 32 if founders and any(term in bg.lower() for term in ELITE_TERMS) else 16 if founders else 0
        theme_pts = min(20, 8 + 5 * len(themes)) if themes else 0
        tech_pts = 18 if any(k in low for k in ["infrastructure", "sdk", "runtime", "cloud vm", "agent", "eval", "robot", "inference", "hardware", "platform", "open source"]) else 10 if "ai" in low else 0
        clarity_pts = min(15, (12 if one_line else 5) + (3 if item.metadata.get("batch") in {"Summer 2026", "Spring 2026", "Winter 2026"} else 0))
        contact_pts = 10 if contacts else 0
        score = founder_pts + theme_pts + tech_pts + clarity_pts + contact_pts
        breakdown = {"founder_background": founder_pts, "theme_fit": theme_pts, "technical_depth": tech_pts, "product_clarity_or_traction": clarity_pts, "contactability": contact_pts}
    elif item.type == "repo":
        stars = int(item.metadata.get("stars") or 0)
        maint = 22 if any(term in bg.lower() for term in ELITE_TERMS) else 14 if owner else 8
        tech = 23 if any(k in low for k in ["runtime", "inference", "benchmark", "sdk", "sandbox", "scheduler", "compiler", "hal", "robot", "edge", "deploy", "coordination", "authorization", "audit"]) else 14 if any(k in low for k in ["ai", "agent", "llm"]) else 0
        adoption = min(20, int(math.log10(max(stars, 1)) * 8) + (4 if stars >= 1000 else 0) + freshness(item.published_at, 8))
        theme = min(15, 4 * len(themes) + (3 if themes else 0))
        commercial = 11 if any(k in low for k in ["enterprise", "platform", "cloud", "sdk", "service", "deploy", "infrastructure", "agents", "robotics", "hardware", "auth"]) else 0
        score = maint + tech + adoption + theme + min(15, commercial + (4 if contacts else 0))
        breakdown = {"maintainer_background": maint, "technical_depth": tech, "adoption_or_momentum": adoption, "theme_fit": theme, "commercialization_and_contactability": min(15, commercial + (4 if contacts else 0))}
    elif item.type == "paper":
        author_pts = 24 if any(term in bg.lower() for term in ELITE_TERMS) else 10 if authors else 0
        novelty = 22 if any(k in low for k in ["novel", "efficient", "benchmark", "state-of-the-art", "low-latency", "on-device", "robot", "agent", "execution", "world model"]) else 12
        path = 16 if any(k in low for k in ["code", "github", "dataset", "benchmark", "deployment", "system", "framework", "simulator"]) else 10 if any(k in low for k in ["robot", "agent", "inference", "on-device"]) else 0
        theme = min(15, 4 * len(themes) + (3 if themes else 0))
        score = author_pts + novelty + path + theme + 3 + freshness(item.published_at, 5)
        breakdown = {"author_or_lab_background": author_pts, "novelty_and_relevance": novelty, "implementation_path": path, "theme_fit": theme, "contactability_or_project_link": 3 + freshness(item.published_at, 5)}
    else:
        points = int(item.metadata.get("points") or 0)
        trace = 25 if item.url and "news.ycombinator.com/item" not in item.url else 8
        detail = 20 if any(k in low for k in ["build", "launch", "open source", "architecture", "robot", "inference", "agent", "infrastructure", "benchmark", "technical"]) else 10
        founder = 14 if any(k in low for k in ["founder", "we built", "i built", "our team", "yc"]) else 4
        engage = min(15, int(math.log10(max(points, 1)) * 7) + freshness(item.published_at, 5))
        theme = min(10, 3 * len(themes) + (2 if themes else 0))
        score = trace + detail + founder + engage + theme
        breakdown = {"traceable_entity_quality": trace, "technical_or_product_detail": detail, "founder_or_background_clue": founder, "engagement_and_freshness": engage, "theme_fit_and_contactability": theme}

    if any(k in low for k in ["pure ai wrapper", "wrapper app", "thin wrapper"]):
        score -= 15
        breakdown["penalty_pure_wrapper"] = -15
    if any(k in low for k in ["series c", "series d", "ipo", "public company"]):
        score -= 12
        breakdown["penalty_too_late_stage"] = -12
    score = max(0, min(100, int(score)))

    people = [clean(f.get("full_name")) for f in founders if f.get("full_name")] or authors[:5] or [clean(owner.get("name") or owner.get("login") or item.metadata.get("author") or item.metadata.get("owner"))]
    founder_or_contact = clean(", ".join([p for p in people if p]) or "抓取到的元数据中未发现")
    if contacts:
        founder_or_contact = clean(founder_or_contact + " | contact: " + ", ".join(contacts[:4]), 450)

    why = "高度匹配 AI infra / agent / robotics 等主题，且有足够公开证据，值得优先人工阅读。" if score >= 80 else "主题相关度不错，公开 artifact 较新或技术描述较具体，适合快速 diligence。" if score >= 70 else "观察项：有上下文价值，但在建联前需要更多证据。"
    next_action = {
        "company": "打开 YC/公司页面，核实 founder profile，并判断是否有 warm intro 或公开邮箱。",
        "repo": "先看 README、issues 和近期 commits，再确认 maintainer 是否有清晰联系方式。",
        "paper": "检查作者主页、代码 release，以及是否存在创业或实验室商业化线索。",
        "post": "打开原帖和外链，识别背后的公司/作者，再决定是否跟踪或联系。",
    }[item.type]
    greeting = founder_or_contact.split("|")[0].strip() or item.name
    english_subject = f"Technical exchange around {item.name}"
    english_body = clean(f"Hi {greeting}, I came across {item.name} through {item.source}. The specific signal that stood out was: {one_line}. We are studying early AI infrastructure, agents, inference, robotics, and cross-border commercialization opportunities. Would you be open to a short technical conversation next week?", 900)
    chinese_subject = f"想就 {item.name} 做一次技术交流"
    chinese_body = clean(f"你好，看到 {item.name} 这个信号，尤其是「{one_line}」很贴近我们近期关注的 AI 基础设施、agent、推理、机器人和跨境技术商业化方向。想先做一次偏技术和产品路线的交流，了解你们正在解决的问题，也分享一些我们在早期市场和资本侧看到的观察。下周方便约 20 分钟聊聊吗？", 900)

    return Scored(
        item=item,
        score=score,
        tier=tier(score),
        outreach_status="draft_ready" if passes_outreach_filter(score, contacts, china) else "not_candidate",
        founder_or_contact=founder_or_contact,
        china_signal=china,
        founder_background_signal=bg,
        theme_tags=themes,
        one_line_summary=one_line,
        evidence=evidence,
        markdown_anchor="",
        why=why,
        next_action=next_action,
        contact_paths=contacts,
        english_subject=english_subject,
        english_body=english_body,
        chinese_subject=chinese_subject,
        chinese_body=chinese_body,
        breakdown=breakdown,
    )


def update_master(rows: list[Scored]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if MASTER_PATH.exists():
        wb = load_workbook(MASTER_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "master"
        ws.append(COLUMNS)
    if [ws.cell(1, i + 1).value for i in range(len(COLUMNS))] != COLUMNS:
        ws = wb.create_sheet("master_current")
        ws.append(COLUMNS)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    existing = {str(ws.cell(r, COLUMNS.index("url") + 1).value): r for r in range(2, ws.max_row + 1) if ws.cell(r, COLUMNS.index("url") + 1).value}
    for row in rows:
        values = {
            "first_seen_date": RUN_DATE,
            "last_seen_date": RUN_DATE,
            "source": row.item.source,
            "type": row.item.type,
            "name_or_title": row.item.name,
            "url": row.item.url,
            "score": row.score,
            "tier": row.tier,
            "outreach_status": row.outreach_status,
            "founder_or_contact": row.founder_or_contact,
            "china_signal": row.china_signal,
            "founder_background_signal": row.founder_background_signal,
            "theme_tags": ", ".join(row.theme_tags),
            "one_line_summary": row.one_line_summary,
            "evidence": " | ".join(row.evidence),
            "markdown_anchor": row.markdown_anchor,
            "notes": "",
        }
        if row.item.url in existing:
            r = existing[row.item.url]
            for col in COLUMNS[1:]:
                if col == "markdown_anchor" and not values[col] and ws.cell(r, COLUMNS.index(col) + 1).value:
                    continue
                ws.cell(r, COLUMNS.index(col) + 1, values[col])
        else:
            ws.append([values[col] for col in COLUMNS])
    for column, width in {"A": 14, "B": 14, "C": 14, "D": 12, "E": 42, "F": 48, "G": 8, "H": 8, "I": 16, "J": 46, "K": 36, "L": 58, "M": 28, "N": 60, "O": 54, "P": 40, "Q": 30}.items():
        ws.column_dimensions[column].width = width
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(MASTER_PATH)


def render_report(top: list[Scored], candidates: list[Scored], coverage: list[dict], failures: list[dict]) -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    lines = [
        f"# {RUN_DATE} AI/科技投资 Sourcing 日报",
        "",
        f"运行时间：{RUN_TS} Asia/Shanghai",
        "",
        "## 覆盖情况",
        "",
        "| 来源 | 状态 | 入选数 | 本轮阈值/说明 |",
        "| --- | --- | ---: | --- |",
    ]
    for c in coverage:
        lines.append(f"| {c['source']} | {c['status']} | {c['count']} | {clean(c.get('notes'), 180)} |")
    if failures:
        lines.extend(["", "### 抓取/解析失败", ""])
        for f in failures:
            lines.append(f"- {f.get('source')} / {f.get('action')}：{clean(f.get('reason'), 220)}（{f.get('status')}；建议重试：{f.get('retry')}）")
    lines.extend(["", "## 今日 Top", ""])
    for idx, row in enumerate(top, 1):
        lines.append(f'<a id="{idx}-{slugify(row.item.name)}"></a>')
        lines.extend(
            [
                f"### {idx}. {row.item.name}",
                "",
                f"- 类型/来源：{row.item.type} / {row.item.source}",
                f"- 分数/层级：{row.score} / {row.tier}",
                f"- 一句话：{row.one_line_summary}",
                f"- 为什么值得看：{row.why}",
                f"- 创始人/作者/维护者背景：{row.founder_background_signal}",
                f"- 中国/华人相关信号：{row.china_signal}",
                f"- 主题标签：{', '.join(row.theme_tags) if row.theme_tags else '未观察到'}",
                f"- 入选阈值证据：{row.item.metadata.get('filter_reason', '未记录')}",
                f"- 证据链接：{' | '.join(row.evidence)}",
                f"- 建议下一步：{row.next_action}",
                f"- 联系路径：{', '.join(row.contact_paths) if row.contact_paths else '抓取到的公开元数据中未发现'}",
                "",
            ]
        )
    lines.extend(["## 建联候选（仅限疑似中国/华人信号）", ""])
    if not candidates:
        lines.append(f"本轮没有同时满足“分数 >= {OUTREACH_POLICY['min_score']} + 有公开联系路径 + 疑似中国/华人信号”的候选。")
    else:
        for idx, row in enumerate(candidates, 1):
            lines.extend(
                [
                    f"{idx}. {row.item.name} ({row.item.source}, score {row.score})",
                    f"   - 中国/华人相关信号：{row.china_signal}",
                    f"   - 联系路径：{', '.join(row.contact_paths)}",
                    f"   - 英文标题：{row.english_subject}",
                    f"   - 英文正文：{row.english_body}",
                    f"   - 中文标题：{row.chinese_subject}",
                    f"   - 中文正文：{row.chinese_body}",
                    f"   - 证据：{' | '.join(row.evidence)}",
                ]
            )
    lines.extend(["", "未发送任何邮件。Gmail 发送仍需要先编号选择，再进行第二次明确确认。"])
    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch and score public investment sourcing signals.")
    parser.add_argument(
        "--artifact-only",
        action="store_true",
        help="Write raw/scored artifacts and update the master table, but leave daily report selection/rendering to the Agent.",
    )
    return parser.parse_args()


def scored_payload(row: Scored) -> dict:
    return {
        "source": row.item.source,
        "type": row.item.type,
        "name": row.item.name,
        "url": row.item.url,
        "score": row.score,
        "tier": row.tier,
        "outreach_status": row.outreach_status,
        "founder_or_contact": row.founder_or_contact,
        "china_signal": row.china_signal,
        "founder_background_signal": row.founder_background_signal,
        "theme_tags": row.theme_tags,
        "one_line_summary": row.one_line_summary,
        "why": row.why,
        "next_action": row.next_action,
        "contact_paths": row.contact_paths,
        "english_subject": row.english_subject,
        "english_body": row.english_body,
        "chinese_subject": row.chinese_subject,
        "chinese_body": row.chinese_body,
        "filter_reason": row.item.metadata.get("filter_reason"),
        "breakdown": row.breakdown,
        "evidence": row.evidence,
    }


def main() -> int:
    args = parse_args()
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    failures = []
    coverage = []
    items = []
    for source, fetcher in [("YC", fetch_yc), ("GitHub", fetch_github), ("arXiv", fetch_arxiv), ("Hacker News", fetch_hn)]:
        before = len(failures)
        try:
            fetched = fetcher(failures)
            items.extend(fetched)
            status = "ok" if fetched else "no_data"
            notes = FILTER_DESCRIPTIONS.get(source, "public source attempted; cap 20")
            if len(failures) > before:
                status = "partial" if fetched else "fetch_failed"
                notes = f"{notes}; {len(failures) - before} failures recorded"
            coverage.append({"source": source, "status": status, "count": len(fetched), "notes": notes})
        except Exception as exc:
            failures.append({"source": source, "action": "source_fetcher", "url": "", "time": RUN_TS, "reason": repr(exc), "status": "fetch_failed", "retry": "inspect source fetcher"})
            coverage.append({"source": source, "status": "fetch_failed", "count": 0, "notes": repr(exc)})
    by_url = {}
    for item in items:
        if item.url and item.url not in by_url:
            by_url[item.url] = item
    scored = [score_item(item) for item in by_url.values()]
    scored.sort(key=lambda row: (row.score, 1 if row.contact_paths else 0, freshness(row.item.published_at, 10)), reverse=True)
    top = scored[:TOP_LIMIT]
    if not args.artifact_only:
        for idx, row in enumerate(top, 1):
            row.markdown_anchor = f"{REPORT_PATH.name}#{idx}-{slugify(row.item.name)}"
        for row in scored[15:]:
            row.markdown_anchor = REPORT_PATH.name
    update_master(scored)
    candidates = [row for row in scored if row.outreach_status == "draft_ready"][: int(OUTREACH_POLICY["max_candidates"])]
    if not args.artifact_only:
        render_report(top, candidates, coverage, failures)
    RAW_PATH.write_text(
        json.dumps(
            {
                "run_date": RUN_DATE,
                "run_timestamp": RUN_TS,
                "mode": "artifact_only" if args.artifact_only else "legacy_report",
                "coverage": coverage,
                "filters": FILTER_DESCRIPTIONS,
                "failures": failures,
                "items": [asdict(item) for item in items],
                "scored": [scored_payload(row) for row in scored],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "run_date": RUN_DATE,
                "coverage": coverage,
                "failures": len(failures),
                "total_unique_items": len(scored),
                "provisional_top_count": len(top),
                "provisional_outreach_candidates": len(candidates),
                "master_path": str(MASTER_PATH),
                "report_path": None if args.artifact_only else str(REPORT_PATH),
                "raw_path": str(RAW_PATH),
                "provisional_top": [{"name": row.item.name, "source": row.item.source, "score": row.score, "tier": row.tier, "url": row.item.url} for row in top],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
